import openpyxl

wb = openpyxl.load_workbook('example.xlsx')

print(wb.sheetnames) # ['Sheet1', 'Sheet2', 'Sheet3']

hoja = wb.active
print(hoja.title) # 'Sheet1'

wb.create_sheet(title='Hoja 4')

